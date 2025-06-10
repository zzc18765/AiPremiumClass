import openpyxl
import matplotlib.pyplot as plt

def read_excel(file_path):
    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)
    
    # 获取所有表名
    sheet_names = workbook.sheetnames
    sheet_names
    
    # 遍历不同工作表
    comment_data = []
    for sheet_name in sheet_names:
        # 获取当前工作表的数据
        sheet = workbook[sheet_name]
        
        # 打印数据
        print(f"\n工作表 '{sheet_name}' 的内容：")

        # 处理数据
        for row in sheet.iter_rows(values_only=True):
            score = row[4]
            content = row[5]
            # 空值过滤
            if score is None or content is None:
                continue
            # 无意义数据过滤
            if content == "此用户未填写评价内容":
                continue
            # 过短或过长的评论过滤
            if len(content) < 5 or len(content) > 150:
                continue
            # 去除换行
            content = content.replace('\n', ' ').replace('\r', ' ')
            comment_data.append((content, score))

        # comment_len = [len(comment[0]) for comment in comment_data]

        # plt.hist(comment_len, bins=20, color='blue', alpha=0.7)
        # plt.xlabel('Comment Length')
        # plt.ylabel('Frequency')
        # plt.title('Comment Length Distribution')
        # plt.show()
        print("评论长度：", len(comment_data))
    
    # 关闭工作簿
    workbook.close()
    
    return comment_data

def save_processed_data(data, file_path):
    with open(file_path, 'w', encoding='utf-8') as f:
        for row in data:
            f.write('\t'.join(map(str, row)) + '\n')

if __name__ == "__main__":
    exl_file_path = "jd_comment_data.xlsx"  # Excel 文件路径
    data = read_excel(exl_file_path)

    save_file_path = "jd_comment_processed.txt"
    save_processed_data(data, save_file_path)